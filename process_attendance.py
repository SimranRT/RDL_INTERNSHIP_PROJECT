import pandas as pd
import openpyxl
from datetime import datetime, timedelta

def process_attendance_excel(file_path):
    # Load the workbook to access multiple sheets
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 1. Extract Employee Info from "Schedule Information Report"
    # Assuming it's the first sheet
    sched_sheet = wb.worksheets[0]
    employee_map = {}
    
    # Data starts from row 5 based on images
    for row in sched_sheet.iter_rows(min_row=5, values_only=True):
        emp_id = row[0]
        name = row[1]
        dept = row[2]
        if emp_id:
            employee_map[str(emp_id)] = {
                'Name': name,
                'Dept': dept
            }
            
    # 2. Extract Attendance Logs from "Att.log report"
    # Assuming it's the third sheet or named "Att.log report"
    log_sheet = None
    for sheet in wb.worksheets:
        if "Att.log" in sheet.title:
            log_sheet = sheet
            break
            
    if not log_sheet:
        return "Error: Att.log report sheet not found"
        
    results = []
    current_emp = None
    
    # Iterate through rows to find ID blocks
    for row_idx, row in enumerate(log_sheet.iter_rows(values_only=True), 1):
        row_val = str(row[0]) if row[0] else ""
        
        if "ID:" in row_val:
            # Found a header row
            # Extract ID from "ID: 1"
            try:
                emp_id = row_val.split("ID:")[1].strip()
                # Name is usually in column J (index 9)
                emp_name = str(row[9]).split("Name:")[1].strip() if row[9] and "Name:" in str(row[9]) else ""
                # Dept is usually in column U (index 20)
                emp_dept = str(row[20]).split("Dept.:")[1].strip() if row[20] and "Dept.:" in str(row[20]) else ""
                
                current_emp = {
                    'id': emp_id,
                    'name': emp_name,
                    'dept': emp_dept
                }
            except Exception as e:
                print(f"Error parsing header at row {row_idx}: {e}")
                continue
        
        elif current_emp and any(row[0:31]):
            # This row likely contains the punch times for days 1-31
            # Check if this row actually has time-like data
            # In the report, the data row is usually right after the header or 1 row below
            
            # We need to determine if this is the data row. 
            # Usually, it's the row where columns 0-30 have content.
            # Let's check if the first cell has a newline or looks like a time
            first_cell = str(row[0]) if row[0] else ""
            if ":" in first_cell:
                # Process this row for days 1-31
                for day in range(1, 32):
                    cell_val = row[day-1] # Days 1-31 are in columns A-AE (0-30)
                    if not cell_val:
                        continue
                        
                    # Split by newline and clean up
                    times = [t.strip() for t in str(cell_val).split('\n') if t.strip()]
                    
                    # Handle biometric double tap (remove duplicates or very close entries)
                    unique_times = []
                    for t in sorted(times):
                        if not unique_times:
                            unique_times.append(t)
                        else:
                            # If difference is less than 2 minutes, ignore
                            t1 = datetime.strptime(unique_times[-1], "%H:%M")
                            t2 = datetime.strptime(t, "%H:%M")
                            if (t2 - t1).total_seconds() > 120:
                                unique_times.append(t)
                    
                    if not unique_times:
                        continue
                        
                    # Map to Login1, Logout1, Login2, Logout2
                    login1 = unique_times[0] if len(unique_times) > 0 else ""
                    logout1 = unique_times[1] if len(unique_times) > 1 else ""
                    login2 = unique_times[2] if len(unique_times) > 2 else ""
                    logout2 = unique_times[-1] if len(unique_times) > 3 else ""
                    
                    # Calculate Overtime
                    overtime = "00:00"
                    try:
                        if login1 and logout2:
                            fmt = "%H:%M"
                            t_login1 = datetime.strptime(login1, fmt)
                            t_logout1 = datetime.strptime(logout1, fmt) if logout1 else None
                            t_login2 = datetime.strptime(login2, fmt) if login2 else None
                            t_logout2 = datetime.strptime(logout2, fmt)
                            
                            total_minutes = 0
                            if t_logout1 and t_login1:
                                total_minutes += (t_logout1 - t_login1).total_seconds() / 60
                            if t_logout2 and t_login2:
                                total_minutes += (t_logout2 - t_login2).total_seconds() / 60
                            elif t_logout2 and t_login1 and not t_logout1:
                                # Single punch pair
                                total_minutes = (t_logout2 - t_login1).total_seconds() / 60
                                
                            if total_minutes > 480: # 8 hours
                                ot_mins = int(total_minutes - 480)
                                overtime = f"{ot_mins // 60:02d}:{ot_mins % 60:02d}"
                    except:
                        pass
                        
                    results.append({
                        'EmpID': current_emp['id'],
                        'Name': current_emp['name'] or employee_map.get(current_emp['id'], {}).get('Name', ''),
                        'Dept': current_emp['dept'] or employee_map.get(current_emp['id'], {}).get('Dept', ''),
                        'Date': f"2025-12-{day:02d}", # Year/Month should be parsed from sheet header
                        'Login1': login1,
                        'Logout1': logout1,
                        'Login2': login2,
                        'Logout2': logout2,
                        'Overtime': overtime
                    })
                
                # After processing the data row, reset current_emp to find the next one
                current_emp = None

    # Convert to DataFrame and save
    df = pd.DataFrame(results)
    df.to_excel("processed_attendance.xlsx", index=False)
    return df

# Example usage:
# process_attendance_excel("attendance_report.xlsx")
